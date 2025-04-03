import os
import json
import time
from datetime import datetime
import pandas as pd
from dotenv import load_dotenv
from tencentcloud.common import credential
from tencentcloud.common.profile.client_profile import ClientProfile
from tencentcloud.common.profile.http_profile import HttpProfile
from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
from tencentcloud.cam.v20190116 import cam_client, models
from openpyxl import load_workbook
from openpyxl.styles import numbers

# 加载环境变量
load_dotenv()

class TencentCloudExporter:
    def __init__(self):
        self.cred = credential.Credential(
            os.getenv("TENCENTCLOUD_SECRET_ID"),
            os.getenv("TENCENTCLOUD_SECRET_KEY")
        )
        self.client = self._init_cam_client()

    def _init_cam_client(self):
        """初始化CAM客户端"""
        http_profile = HttpProfile()
        http_profile.endpoint = "cam.tencentcloudapi.com"
        client_profile = ClientProfile()
        client_profile.httpProfile = http_profile
        return cam_client.CamClient(self.cred, "", client_profile)

    # ---------------------- 用户数据获取 ----------------------
    def _process_users(self, users_data):
        """处理子用户数据结构"""
        return [{
            "用户名称": u.get("Name", "N/A"),
            "用户类型": "子用户",
            "账号ID": str(u.get("Uin", "")),
            "备注信息": u.get("Remark", ""),
            "控制台登录": "允许" if u.get("ConsoleLogin") else "禁止"
        } for u in users_data]

    def get_all_users(self):
        """获取所有子用户"""
        try:
            req = models.ListUsersRequest()
            params = {
            }
            req.from_json_string(json.dumps(params))
            resp = self.client.ListUsers(req)
            data = json.loads(resp.to_json_string())
            users = data.get("Data", [])
            return self._process_users(users)
        except TencentCloudSDKException as e:
            print(f"获取所有子用户失败: {e}")

    def _process_collaborators(self, collaborators_data):
        """处理协作者数据结构"""
        return [{
            "用户名称": u.get("Name", "N/A"),
            "用户类型": "协作者",
            "账号ID": str(u.get("Uin", "")),
            "备注信息": u.get("Remark", ""),
            "控制台登录": "允许" if u.get("ConsoleLogin") else "禁止"
        } for u in collaborators_data]

    def get_all_collaborators(self):
        """获取所有协作者"""
        try:
            req = models.ListCollaboratorsRequest()
            resp = self.client.ListCollaborators(req)
            data = json.loads(resp.to_json_string())
            users = data.get("Data", [])
            return self._process_collaborators(users)
        except TencentCloudSDKException as e:
            print(f"获取所有协作者失败: {e}")
            return []


    # ---------------------- 策略数据获取 ----------------------
    def get_all_policies(self):
        """分页获取所有策略及关联用户"""
        policies = []
        page = 0
        rp = 200  # 每页策略数量

        try:
            # 阶段1：获取所有策略基础信息
            while True:
                time.sleep(0.1)
                page += 1
                req = models.ListPoliciesRequest()
                req.Page = page
                req.Rp = rp
                resp = self.client.ListPolicies(req)
                data = json.loads(resp.to_json_string())
                batch = data.get("List", [])

                # 阶段2：遍历每个策略获取关联用户
                for policy in batch:
                    policy_id = int(policy.get("PolicyId"))
                    users = []
                    entity_page = 0

                    # 分页获取关联实体
                    while True:
                        time.sleep(0.02)
                        entity_page += 1
                        req = models.ListEntitiesForPolicyRequest()
                        params = {
                            "PolicyId": policy_id,
                            "Page": entity_page,
                            "Rp": rp,
                            "EntityFilter": "User"  # 仅获取用户类型实体
                        }
                        req.from_json_string(json.dumps(params))

                        resp = self.client.ListEntitiesForPolicy(req)
                        entity_data = json.loads(resp.to_json_string())

                        # 过滤用户类型实体（RelatedType=1）
                        users.extend([
                            {
                                "账号ID": str(e.get("Uin", "")),
                                "用户名称": str(e.get("Name", "")),
                                "关联时间": str(e.get("AttachmentTime", ""))
                            }
                            for e in entity_data.get("List", [])
                            if e.get("RelatedType") == 1
                        ])

                        if len(entity_data.get("List", [])) < rp:
                            break

                    # 构造策略数据结构
                    policies.append({
                        "策略名称": policy.get("PolicyName", "N/A"),
                        "策略类型": "预设" if policy.get("Type") == 2 else "自定义",
                        "策略描述": policy.get("Description", ""),
                        "关联用户": users  # 存储用户ID及关联类型
                    })

                if len(batch) < rp:
                    break

        except TencentCloudSDKException as e:
            print(f"策略查询失败: {e}")

        return policies

    # ---------------------- Excel导出逻辑 ----------------------
    def export_accounts(self):
        filename = f"{datetime.now().year}年度腾讯云账号权限清单.xlsx"
        with pd.ExcelWriter(
                filename,
                engine='openpyxl',
                mode='w'
        ) as writer:
            # 初始化 Sheet
            pd.DataFrame(columns=["用户名称", "用户类型", "账号ID", "备注信息", "控制台登录"]).to_excel(
                writer,
                sheet_name='用户清单',
                index=False
            )
            pd.DataFrame(columns=["策略名称", "策略类型", "策略描述"]).to_excel(
                writer,
                sheet_name='策略清单',
                index=False
            )
            pd.DataFrame(columns=["用户名称", "账号ID", "策略名称", "策略描述"]).to_excel(
                writer,
                sheet_name='策略关联',
                index=False
            )

            # 获取所有数据
            users = self.get_all_users()
            collaborators = self.get_all_collaborators()
            combined_users = users + collaborators  # 合并子用户和协作者数据
            policies = self.get_all_policies()

            # 用户清单写入
            if combined_users:
                df_users = pd.DataFrame(combined_users)
                df_users.to_excel(writer, sheet_name='用户清单', index=False)
                self._format_sheet(writer, '用户清单', {
                    'A': 20, 'B': 12, 'C': 18, 'D': 30, 'E': 15
                })

            # 策略清单写入（排除关联用户）
            if policies:
                df_policies = pd.DataFrame([
                    {
                        "策略名称": p["策略名称"],
                        "策略类型": p["策略类型"],
                        "策略描述": p["策略描述"]
                    } for p in policies
                ])
                df_policies.to_excel(writer, sheet_name='策略清单', index=False)
                self._format_sheet(writer, '策略清单', {'A': 30, 'B': 12, 'C': 150})

            # 策略关联处理
            relations = []
            for policy in policies:
                for user_info in policy.get("关联用户", []):
                    relations.append({
                        "用户名称": user_info.get("用户名称", "N/A"),
                        "账号ID": str(user_info.get("账号ID", "")),
                        "策略名称": policy["策略名称"],
                        "策略描述": policy["策略描述"]
                    })

            if relations:
                df_relations = pd.DataFrame(relations)

                # 按名称排序
                df_sorted = df_relations.sort_values(by="用户名称").reset_index(drop=True)
                df_sorted.to_excel(writer, sheet_name='策略关联', index=False)
                df_relations.to_excel(writer, sheet_name='策略关联', index=False)

                # 应用格式设置
                self._format_sheet(writer, '策略关联', {
                    'A': 20, 'B': 20, 'C': 30, 'D': 150
                })

        self._post_process_excel(filename)
        print(f"文件已生成：{filename}")

    def _format_sheet(self, writer, sheet_name, widths=None):
        """通用表格格式化"""
        workbook = writer.book
        ws = workbook[sheet_name]

        # 设置列宽
        for col, width in (widths or {}).items():
            ws.column_dimensions[col].width = width

        # 首行冻结
        ws.freeze_panes = 'A2'

        # 处理科学计数法
        if sheet_name == '用户清单':
            for col in ['C']:  # 账号ID列
                for cell in ws[col]:
                    cell.number_format = numbers.FORMAT_TEXT

        if sheet_name == '策略关联':
            for col in ['B']:  # 账号ID
                for cell in ws[col]:
                    cell.number_format = numbers.FORMAT_TEXT

    def _post_process_excel(self, filename):
        """最终格式优化"""
        wb = load_workbook(filename)

        # 删除空Sheet
        for sheet in ['Sheet1', 'Sheet']:
            if sheet in wb.sheetnames and wb[sheet].max_row == 1:
                del wb[sheet]

        # 设置默认视图
        wb.active = wb['用户清单']
        wb.save(filename)


if __name__ == "__main__":
    try:
        exporter = TencentCloudExporter()
        exporter.export_accounts()
        print("导出成功，文件已生成")
    except Exception as e:
        print(f"执行异常: {str(e)}")