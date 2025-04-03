import json
import os

import pandas as pd
from datetime import datetime
from io import BytesIO
from curl_cffi import requests as cffi_requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 初始配置
load_dotenv()

headers = {
    "PRIVATE-TOKEN": os.getenv("ACCESS_TOKEN")
}
current_year = datetime.now().year
path = f'{current_year}年度腾讯工蜂Git权限清单.xlsx'

# 访问权限映射
access_levels = {
    10: ('Guest', '游客'),
    15: ('Follower', '允许浏览代码'),
    20: ('Reporter', '允许下载代码'),
    30: ('Developer', '允许读写代码'),
    40: ('Master', '允许管理项目和代码'),
    50: ('Owner', '允许管理仓库、项目和代码')
}

# 需要获取的项目组列表
target_groups = ['pyfund', 'py-components', 'pyadmin', 'puyi-app', 'tougu', 'pyorg', 'pay']
subgroups_of_fundtrade = {'pay'}
fundtrade_path = 'fundtrade'

# 获取项目组
def fetch_group_names():
    group_url = "https://git.code.tencent.com/api/v3/groups"
    params = {
        "page": 1,
        "per_page": 100
    }
    try:
        group_response = cffi_requests.get(group_url, params=params, headers=headers, verify=True)
        group_response.raise_for_status()
        groups_data = json.loads(group_response.content.decode())
        # 过滤出目标项目组及其子组
        filtered_groups = [group for group in groups_data if group['path'] in target_groups]

        # 获取 fundtrade 的子组并过滤出目标子组
        fundtrade_id = next((group['id'] for group in groups_data if group['path'] == fundtrade_path), None)
        if fundtrade_id:
            subgroups = fetch_subgroups(fundtrade_id)
            filtered_groups.extend(subgroup for subgroup in subgroups if subgroup['path'] in subgroups_of_fundtrade)

        return filtered_groups
    except Exception as e:
        print(f"请求或处理数据时发生错误: {e}")
        return []


def fetch_subgroups(parent_id):
    subgroup_url = f"https://git.code.tencent.com/api/v3/groups/{parent_id}/subgroups"
    try:
        response = cffi_requests.get(subgroup_url, headers=headers, verify=True)
        response.raise_for_status()
        subgroups_data = json.loads(response.content.decode())
        return subgroups_data
    except Exception as e:
        print(f"获取子组信息时发生错误: {e}")
        return []


# 获取项目组的详细信息
def fetch_group_details(group_id):
    group_details_url = f"https://git.code.tencent.com/api/v3/groups/{group_id}"
    try:
        response = cffi_requests.get(group_details_url, headers=headers, verify=True)
        response.raise_for_status()
        group_details = json.loads(response.content.decode())
        return group_details
    except Exception as e:
        print(f"获取项目组 {group_id} 详细信息时发生错误: {e}")
        return None


# 处理单个成员的信息，并增加访问权限说明
def process_member(member):
    username = member.get('username')
    name = member.get('name')
    access_level_num = member.get('access_level')
    access_level, description = access_levels.get(access_level_num, ('未知', '未知权限'))
    state = '正常' if member.get('state') == 'active' else member.get('state', '未知')
    return [username, name, state, access_level, description]


# 创建一个Workbook对象，允许写入多个sheet
wb = Workbook()
ws_project_info = wb.active
ws_project_info.title = '项目组与项目信息'

# 写入表头
ws_project_info.append(['项目组名称', '项目组描述', '项目名称', '项目描述', '项目路径'])

# 第一个sheet写入项目组详细信息
group_details_list = []
groups_data = fetch_group_names()

if not groups_data:
    print("没有找到任何项目组。")
else:
    for group in groups_data:
        group_id = group['id']
        group_details = fetch_group_details(group_id)
        if group_details:
            for project in group_details.get('projects', []):
                group_details_list.append([
                    group_details['name'],
                    group_details['description'],
                    project['name'],
                    project['description'],
                    project['web_url']
                ])

    group_details_df = pd.DataFrame(group_details_list, columns=[
        '项目组名称', '项目组描述', '项目名称', '项目描述', '项目路径'
    ])

    # 将 DataFrame 的每一行转换为列表并追加到工作表中
    for row in group_details_df.itertuples(index=False, name=None):
        ws_project_info.append(row)

    print("项目组与项目信息已成功保存至Excel.")

    # 处理成员信息
    processed_groups = set()  # 用于跟踪已经处理过的组
    for group in groups_data:
        group_name = group['path']
        parent_group_name = ''
        if group_name in subgroups_of_fundtrade:
            parent_group_name = fundtrade_path

        full_group_name = f"{parent_group_name}/{group_name}" if parent_group_name else group_name
        if full_group_name in processed_groups:
            continue  # 如果已经处理过，则跳过

        if parent_group_name:
            url = f"https://git.code.tencent.com/api/v3/groups/{parent_group_name}%2f{group_name}/members"
        else:
            url = f"https://git.code.tencent.com/api/v3/groups/{group_name}/members"

        print(f"正在处理组：{full_group_name}, URL: {url}")
        try:
            response = cffi_requests.get(url, headers=headers, verify=True)
            response.raise_for_status()
            content = BytesIO(response.content)
            members_data = json.load(content)
            processed_members = [process_member(member) for member in members_data]
            df = pd.DataFrame(processed_members, columns=['用户名', '昵称', '状态', '访问权限', '说明'])

            if not df.empty:
                ws_members = wb.create_sheet(title=group_name[:31])  # Sheet 名称最多31个字符
                # 写入表头
                ws_members.append(['用户名', '昵称', '状态', '访问权限', '说明'])

                # 将 DataFrame 的每一行转换为列表并追加到工作表中
                for row in df.itertuples(index=False, name=None):
                    ws_members.append(row)
                print(f"{group_name} 的数据已成功保存至Excel.")
            else:
                print(f"{group_name} 没有任何成员信息。")
        except cffi_requests.HTTPError as http_err:
            if http_err.response.status_code == 403:
                print(f"处理 {group_name} 时发生错误：HTTP Error 403: Forbidden")
            else:
                print(f"处理 {group_name} 时发生 HTTP 错误: {http_err}")
        except Exception as e:
            print(f"处理 {group_name} 时发生错误：{e}")

        processed_groups.add(full_group_name)  # 标记为已处理

# 调整列宽并合并单元格
for sheet in wb.worksheets:
    # 合并项目组名称和项目组描述的单元格
    if sheet.title == '项目组与项目信息':
        start_row = 2  # 数据从第2行开始（第1行是表头）
        current_group_name = None
        merge_start_row = start_row

        for row in range(start_row, sheet.max_row + 1):
            group_name = sheet.cell(row=row, column=1).value  # 项目组名称在第1列
            if group_name != current_group_name:
                if current_group_name is not None:
                    # 合并上一个项目组的单元格
                    sheet.merge_cells(start_row=merge_start_row, end_row=row - 1, start_column=1,
                                      end_column=1)  # 合并项目组名称
                    sheet.merge_cells(start_row=merge_start_row, end_row=row - 1, start_column=2,
                                      end_column=2)  # 合并项目组描述
                current_group_name = group_name
                merge_start_row = row

        # 合并最后一个项目组的单元格
        if current_group_name is not None:
            sheet.merge_cells(start_row=merge_start_row, end_row=sheet.max_row, start_column=1, end_column=1)  # 合并项目组名称
            sheet.merge_cells(start_row=merge_start_row, end_row=sheet.max_row, start_column=2, end_column=2)  # 合并项目组描述

    # 调整列宽
    for col in sheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 15)
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

wb.save(path)
print("所有操作完成。")


