import json
import os

import pandas as pd
from datetime import datetime
from io import BytesIO
from curl_cffi import requests as cffi_requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import time
import logging

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 初始配置
load_dotenv()

headers = {
    "PRIVATE-TOKEN": os.getenv("ACCESS_TOKEN")
}
current_year = datetime.now().year
path = f'{current_year}年度腾讯工蜂Git权限清单.xlsx'

# 访问权限映射
access_levels = {
    10: ('Guest', '游客'),
    15: ('Follower', '允许浏览代码'),
    20: ('Reporter', '允许下载代码'),
    30: ('Developer', '允许读写代码'),
    40: ('Master', '允许管理项目和代码'),
    50: ('Owner', '允许管理仓库、项目和代码')
}

# 获取项目组
def fetch_group_names():
    group_url = "https://git.code.tencent.com/api/v3/groups"
    params = {
        "page": 1,
        "per_page": 100
    }
    group_names = []
    groups_data = []
    while True:
        try:
            group_response = cffi_requests.get(group_url, params=params, headers=headers, verify=True)
            group_response.raise_for_status()
            groups_data_page = json.loads(group_response.content.decode())
            group_names.extend([group['path'] for group in groups_data_page])
            groups_data.extend(groups_data_page)
            if len(groups_data_page) < params["per_page"]:
                break
            params["page"] += 1
            time.sleep(1)  # 避免触发速率限制
        except Exception as e:
            logging.error(f"请求或处理数据时发生错误: {e}")
            break
    return group_names, groups_data

group_names, groups_data = fetch_group_names()

# 获取项目组的详细信息
def fetch_group_details(group_id):
    group_details_url = f"https://git.code.tencent.com/api/v3/groups/{group_id}"
    try:
        response = cffi_requests.get(group_details_url, headers=headers, verify=True)
        response.raise_for_status()
        group_details = json.loads(response.content.decode())
        return group_details
    except Exception as e:
        logging.error(f"获取项目组 {group_id} 详细信息时发生错误: {e}")
        return None

# 处理单个成员的信息，并增加访问权限说明
def process_member(member):
    username = member.get('username')
    name = member.get('name')
    access_level_num = member.get('access_level')
    access_level, description = access_levels.get(access_level_num, ('未知', '未知权限'))
    state = '正常' if member.get('state') == 'active' else member.get('state', '未知')
    return [username, name, state, access_level, description]

# 创建一个ExcelWriter对象，允许写入多个sheet
with pd.ExcelWriter(path, engine='openpyxl', mode='w') as writer:
    # 第一个sheet写入项目组详细信息
    group_details_list = []
    for group in groups_data:
        group_id = group['id']
        group_details = fetch_group_details(group_id)
        if group_details:
            for project in group_details.get('projects', []):
                group_details_list.append([
                    group_details['name'],
                    group_details['description'],
                    project['name'],
                    project['description'],
                    project['web_url']
                ])

    group_details_df = pd.DataFrame(group_details_list, columns=[
        '项目组名称', '项目组描述', '项目名称', '项目描述', '项目路径'
    ])
    group_details_df.to_excel(writer, sheet_name='项目组与项目信息', index=False)
    logging.info("项目组与项目信息已成功保存至Excel.")

    # 处理成员信息
    member_url = "https://git.code.tencent.com/api/v3/groups/{}/members"
    for group_name in group_names:
        url = member_url.format(group_name)
        logging.info(f"正在处理组：{group_name}")
        try:
            response = cffi_requests.get(url, headers=headers, verify=True)
            response.raise_for_status()
            content = BytesIO(response.content)
            members_data = json.load(content)
            processed_members = [process_member(member) for member in members_data]
            df = pd.DataFrame(processed_members, columns=['用户名', '昵称', '状态', '访问权限', '说明'])
            df.to_excel(writer, sheet_name=group_name, index=False)
            logging.info(f"{group_name} 的数据已成功保存至Excel.")
            time.sleep(1)  # 避免触发速率限制
        except Exception as e:
            logging.error(f"处理 {group_name} 时发生错误：{e}")

# 调整列宽并合并单元格
workbook = load_workbook(path)
sheet = workbook['项目组与项目信息']

# 合并项目组名称和项目组描述的单元格
start_row = 2  # 数据从第2行开始（第1行是表头）
current_group_name = None
merge_start_row = start_row

for row in range(start_row, sheet.max_row + 1):
    group_name = sheet.cell(row=row, column=1).value  # 项目组名称在第1列
    if group_name != current_group_name:
        if current_group_name is not None:
            # 合并上一个项目组的单元格
            sheet.merge_cells(start_row=merge_start_row, end_row=row - 1, start_column=1, end_column=1)  # 合并项目组名称
            sheet.merge_cells(start_row=merge_start_row, end_row=row - 1, start_column=2, end_column=2)  # 合并项目组描述
        current_group_name = group_name
        merge_start_row = row

# 合并最后一个项目组的单元格
if current_group_name is not None:
    sheet.merge_cells(start_row=merge_start_row, end_row=sheet.max_row, start_column=1, end_column=1)  # 合并项目组名称
    sheet.merge_cells(start_row=merge_start_row, end_row=sheet.max_row, start_column=2, end_column=2)  # 合并项目组描述

# 调整列宽
for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 15)
        worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

workbook.save(path)
logging.info("所有操作完成。")